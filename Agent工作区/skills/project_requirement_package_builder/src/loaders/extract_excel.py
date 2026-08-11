# -*- coding: utf-8 -*-
"""Excel 抽取：每个 sheet 转 Markdown 表格（处理合并单元格），导出内嵌图片并带锚点。"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from .base import (KIND_IMAGE, KIND_TABLE, FileExtraction, Unit,
                   rows_to_markdown_table, sanitize_name)


def _grid_with_merges(ws, max_rows: int, max_cols: int, fx: FileExtraction):
    """取单元格值网格；合并单元格把左上值标注为「值 ⟨合并RxC⟩」。"""
    nrows = min(ws.max_row or 0, max_rows)
    ncols = min(ws.max_column or 0, max_cols)
    grid = [[ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
            for r in range(1, nrows + 1)]
    try:
        for m in ws.merged_cells.ranges:
            r0, c0 = m.min_row - 1, m.min_col - 1
            if r0 < nrows and c0 < ncols and grid[r0][c0] not in (None, ""):
                span_r, span_c = m.max_row - m.min_row + 1, m.max_col - m.min_col + 1
                if span_r > 1 or span_c > 1:
                    grid[r0][c0] = f"{grid[r0][c0]} ⟨合并{span_r}x{span_c}⟩"
    except Exception as e:  # noqa: BLE001
        fx.notes.append(f"sheet {ws.title}: 合并单元格标注失败（{e}）")
    if (ws.max_row or 0) > max_rows:
        fx.notes.append(f"sheet {ws.title}: 行数 {ws.max_row} 超过上限 {max_rows}，已截断")
    if (ws.max_column or 0) > max_cols:
        fx.notes.append(f"sheet {ws.title}: 列数 {ws.max_column} 超过上限 {max_cols}，已截断")
    return grid


def _trim_grid(grid):
    """去掉全空行与全空尾列。"""
    grid = [r for r in grid if any(c not in (None, "") for c in r)]
    if not grid:
        return grid
    last = max(max((i for i, c in enumerate(r) if c not in (None, "")), default=-1)
               for r in grid)
    return [r[: last + 1] for r in grid]


def extract_excel(path: Path, fx: FileExtraction, assets_dir: Path, cfg: dict) -> None:
    max_rows = cfg.get("excel_max_rows_per_sheet", 400)
    max_cols = cfg.get("excel_max_cols_per_sheet", 40)
    min_bytes = cfg.get("min_image_bytes", 3000)
    stem = sanitize_name(path.stem, 40)

    # data_only=True 读公式缓存值；需要图片锚点所以不能 read_only
    wb = openpyxl.load_workbook(str(path), data_only=True)
    try:
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                fx.notes.append(f"sheet {ws.title}: 非可见（{ws.sheet_state}），仍尝试抽取")
            grid = _trim_grid(_grid_with_merges(ws, max_rows, max_cols, fx))
            if grid:
                md = rows_to_markdown_table(grid)
                fx.add(Unit(locator=f"sheet {ws.title}", kind=KIND_TABLE, text=md,
                            context=f"工作表「{ws.title}」，原始尺寸 {ws.max_row}行x{ws.max_column}列"))

            # 内嵌图片（openpyxl 私有 API，尽力而为）
            for ii, img in enumerate(getattr(ws, "_images", []), start=1):
                try:
                    data = img._data()
                    if len(data) < min_bytes:
                        continue
                    anchor = ""
                    try:
                        frm = img.anchor._from
                        anchor = f"R{frm.row + 1}C{frm.col + 1}"
                    except Exception:  # noqa: BLE001
                        pass
                    fmt = (getattr(img, "format", "") or "png").lower()
                    name = f"{fx.file_id}_{stem}_{sanitize_name(ws.title, 20)}_img{ii}.{fmt}"
                    out = assets_dir / "images" / name
                    out.parent.mkdir(parents=True, exist_ok=True)
                    out.write_bytes(data)
                    fx.add(Unit(locator=f"sheet {ws.title} 图片{ii}（锚点{anchor or '未知'}）",
                                kind=KIND_IMAGE,
                                asset_path=f"assets/images/{name}",
                                needs_visual_reading=True,
                                context=f"工作表「{ws.title}」内嵌图片，锚点单元格 {anchor or '未知'}"))
                except Exception as e:  # noqa: BLE001
                    fx.notes.append(f"sheet {ws.title} 图片{ii}: 导出失败（{e}）")

            if getattr(ws, "_charts", None):
                fx.notes.append(f"sheet {ws.title}: 含 {len(ws._charts)} 个图表，当前不解析图表（预留扩展点）")
    finally:
        wb.close()
